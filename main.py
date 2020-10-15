import pandas as pd
import re
from ratecard import *
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from math import ceil

machines = pd.read_csv('TCO reference - C-Machines.csv')
rates = pd.read_csv('TCO reference - O-Rates.csv')

# csv_to_read = input('path to csv file: ')
# df = pd.read_csv(csv_to_read)

#provisional test on Alphasense
df = pd.read_csv('bill_examples/axcient_rowitems.csv')

#function to detect column names
def detect_column_names(dataframe):
  usage_options = ['lineItem/UsageType', 'UsageType']
  cost_options = ['lineItem/UnblendedCost', 'TotalCost']
  rate_options = ['BlendedRate', 'UnblendedRate', 'lineItem/UnblendedRate']
  description_options = ['lineItem/LineItemDescription', 'ItemDescription']
  quantity_options = ['lineItem/UsageAmount', 'UsageQuantity']
  productname_options = ['product/ProductName', 'ProductCode']
  type_options = ['ItemType']

  columnnames = {
  'usage': list(set(usage_options).intersection(set(dataframe.columns)))[0],
  'cost': list(set(cost_options).intersection(set(dataframe.columns)))[0],
  'rate': list(set(rate_options).intersection(set(dataframe.columns)))[0],
  'description': list(set(description_options).intersection(set(dataframe.columns)))[0],
  'quantity': list(set(quantity_options).intersection(set(dataframe.columns)))[0],
  'productname': list(set(productname_options).intersection(set(dataframe.columns)))[0],
  # 'type': set(type_options).intersection(set(dataframe.columns)))[0],
  }
  return columnnames


#Main component
def sud(nunits, rate):
  n30, remaining = int(nunits)//744, int(nunits)%744
  n15, remaining = remaining//360, remaining%360
  h30, h15, h0 = n30*744, n15*360, remaining
  final = (0.7*h30/nunits + 0.85*h15/nunits + remaining/nunits)*rate
  return final#Returns normalized rate after sud

ratedict = {
    'box': 'List-price',
    'heavy': 'Commitment-1yr',
    'spot': 'Preemptible-use'
}

ssd_rate = {
    'box' : 0.08,
    'heavy' : 0.051,
    'spot' : 0.048
}

gpu_rates = {
    'NVIDIA Tesla K80':{
        'box' : 0.45,
        'heavy' : 0.283, 
        'spot' : 0.135
    },
    'NVIDIA Tesla V100':{
        'box' : 2.48,
        'heavy' : 1.562,
        'spot' : 0.74,
    },
    # 'NVIDIA GRID K520':{
    #     'box' : 2.48,
    #     'heavy' : 1.562,
    #     'spot' : 0.74,
    # },
    # 'NVIDIA Tesla M60':{
    #     'box' : 2.48,
    #     'heavy' : 1.562,
    #     'spot' : 0.74,
    # },
    # 'NVIDIA T4 Tensor Core':{
    #     'box' : 2.48,
    #     'heavy' : 1.562,
    #     'spot' : 0.74,
    # }
}


testmode = False

sud_dict = {
    'GPU' : True,
    'All' : True,
}

def get_usagetype(row):

  val = row[columnnames['usage']]

  if 'BoxUsage' in val:
    usagetype = 'box'
  elif 'HeavyUsage' in val:
    usagetype = 'heavy'
  elif 'SpotUsage' in val:
    usagetype = 'spot'
  # print(val)

  if row[columnnames['productname']]=='Amazon Elastic MapReduce':
    usagetype = 'spot'

  return usagetype

def parsecompute(row):
  val = row[columnnames['usage']]
  usagetype = get_usagetype(row)
  try:
    region, rest = val.split('-')
  except:
    region, rest = 'USW2', val
    # return pd.Series([0]*8)
  try:
    usage, machine = rest.split(':')
  except:
    usage, machine = rest, ''
    return pd.Series([0]*8)
  machine = machine.replace('.elasticsearch', '')
  gcpmachine = machines[machines['API Name']==machine].iloc[0]
  gcpmachine_name = [gcpmachine['Family'].upper(), gcpmachine['Type'].capitalize(), gcpmachine['Memory'], gcpmachine['vCPUs']]
  if gcpmachine_name[1]=='Standard' or gcpmachine_name[1]=='Highmem' or gcpmachine_name[1]=='Highcpu':
    gcpmachine_name[1] = 'Predefined'
  #Needs restructuring to accommodate for all regions and pull rate from the new JSON
  if gcpmachine_name[1] not in ['Small', 'Micro', 'Medium']:
    rates_cpu = rates[(rates['Family']==gcpmachine_name[0])&(rates['Type']==gcpmachine_name[1])&(rates['Metric']=='vCPUs')&(rates['Code']==region_dict[region])].iloc[0][ratedict[usagetype]].replace('$', '')
    rates_mem = rates[(rates['Family']==gcpmachine_name[0])&(rates['Type']==gcpmachine_name[1])&(rates['Metric']=='Memory')&(rates['Code']==region_dict[region])].iloc[0][ratedict[usagetype]].replace('$', '')
    gcp_rate = gcpmachine_name[2]*float(rates_mem)+gcpmachine_name[3]*float(rates_cpu)
    if machine.startswith('p'):
      if gcpmachine['GPU model']=='NVIDIA Tesla K80':
        gcp_rate+=int(gcpmachine['GPUs'])*gpu_rates['NVIDIA Tesla K80'][usagetype]
      elif gcpmachine['GPU model']=='NVIDIA Tesla V100':
        gcp_rate+=int(gcpmachine['GPUs'])*gpu_rates['NVIDIA Tesla V100'][usagetype]
  else:
    gcp_rate = float(rates[(rates['Family']==gcpmachine_name[0])&(rates['Type']==gcpmachine_name[1])].iloc[0][ratedict[usagetype]].replace('$', ''))

  original_rate = gcp_rate
  if row[columnnames['rate']]:
    nunits = row[columnnames['cost']]/row[columnnames['rate']]
    if usagetype=='heavy' or usagetype=='box':
      gcp_rate = sud(nunits, gcp_rate)
  else:
    nunits = 0
  ssd_cost = 0
  if gcpmachine['Instance Storage']!=0 and gcpmachine['SSD/HDD?']=='SSD':
    ssd_cost +=int(gcpmachine['Instance Storage'])*ssd_rate[usagetype]*(nunits//744)
  sud_savings = nunits*(original_rate-gcp_rate)
  # print(ssd_cost, int(gcpmachine['Instance Storage']), ssd_rate[usagetype], nunits, nunits//744)
  # print(nunits, original_rate)
  returner = [nunits]+gcpmachine_name[:2]+[gcp_rate, nunits*gcp_rate+ssd_cost, sud_savings, ssd_cost, sud_savings/(nunits*original_rate)*100]
  return pd.Series(returner)

columnnames = detect_column_names(df)

# for key in columnnames:
# 	columnnames[key] = input('Enter column name for {}: '.format(key))

df = df[df[columnnames['cost']]>0]
df = df[df[columnnames['usage']]!='']
# df = df[~df[columnnames['description']].str.contains('Total')]

df[columnnames['rate']] = df[columnnames['cost']]/df[columnnames['quantity']]
grouped = df.groupby([columnnames['usage'], columnnames['description'], columnnames['productname']]).agg({columnnames['cost']:'sum', columnnames['rate']:'mean', columnnames['quantity']:'sum'}).reset_index()

aws_actual_cost = sum(grouped[columnnames['cost']])
print(aws_actual_cost)

box_filter = grouped[columnnames['usage']].str.contains('BoxUsage')
boxusage = grouped[box_filter].sort_values(columnnames['usage'])
grouped = grouped[~box_filter]

heavy_filter = (grouped[columnnames['usage']].str.contains('HeavyUsage'))&(~grouped[columnnames['usage']].str.contains('HeavyUsage:db')&(grouped[columnnames['productname']]=='AmazonEC2'))
heavyusage = grouped[heavy_filter].sort_values(columnnames['usage'])
grouped = grouped[~heavy_filter]

spot_filter = grouped[columnnames['usage']].str.contains('SpotUsage')
spotusage = grouped[spot_filter].sort_values(columnnames['usage'])
grouped = grouped[~spot_filter]

if len(boxusage):
  print('Box AWS: ', sum(boxusage[columnnames['cost']]))
  boxusage[['nunits', 'gcp_family', 'gcp_type', 'gcp_rate', 'gcp_cost', 'sud_savings', 'ssd_cost', 'sud_savings_percentage']] = boxusage.apply(parsecompute, axis=1)
  # print('Boxusage', sum(boxusage[columnnames['cost']]), sum(boxusage['gcp_cost']))
if len(heavyusage):
  print('heavy AWS: ', sum(heavyusage[columnnames['cost']]), len(heavyusage))
  heavyusage[['nunits', 'gcp_family', 'gcp_type', 'gcp_rate', 'gcp_cost', 'sud_savings', 'ssd_cost', 'sud_savings_percentage']] = heavyusage.apply(parsecompute, axis=1)
  # print('Heavyusage', sum(heavyusage[columnnames['cost']]), sum(heavyusage['gcp_cost']))
if len(spotusage):
  print('spot AWS: ', sum(spotusage[columnnames['cost']]))
  spotusage[['nunits', 'gcp_family', 'gcp_type', 'gcp_rate', 'gcp_cost', 'sud_savings', 'ssd_cost', 'sud_savings_percentage']] = spotusage.apply(parsecompute, axis=1)
  # print('Spotusage', sum(spotusage[columnnames['cost']]), sum(spotusage['gcp_cost']))

######################################################################################################################################################

pd_filter = (grouped[columnnames['usage']].str.contains('VolumeUsage.gp2')) | (grouped[columnnames['usage']].str.contains('SnapshotUsage')) | ((grouped[columnnames['usage']].str.contains('VolumeUsage'))&(grouped[columnnames['description']].str.contains('Magnetic provisioned storage')))
persistentdisk = grouped[pd_filter].sort_values(columnnames['usage'])
grouped = grouped[~pd_filter]

def parsepd(row):
  value = row[columnnames['usage']]
  if not value.startswith('EBS'):
    region, cat = value.split('-EBS:')
  else:
    region, cat = 'USE1', value.replace('EBS:', '')
  if cat=='SnapshotUsage':
    return pd.Series([persistentdisk_ratecard['snapshot'][region_dict[region]], row[columnnames['quantity']]*persistentdisk_ratecard['snapshot'][region_dict[region]]])
  elif cat=='VolumeUsage.gp2':
    return pd.Series([persistentdisk_ratecard['gp2'][region_dict[region]], row[columnnames['quantity']]*persistentdisk_ratecard['gp2'][region_dict[region]]])
  elif 'Magnetic provisioned storage' in row[columnnames['description']]:
    return pd.Series([persistentdisk_ratecard['magnetic'][region_dict[region]], row[columnnames['quantity']]*persistentdisk_ratecard['magnetic'][region_dict[region]]])


persistentdisk[['GCP_rate', 'GCP_cost']] = persistentdisk.apply(parsepd, axis=1)

print("PD AWS: ", sum(persistentdisk[columnnames['cost']]), " PD GCP: ", sum(persistentdisk['GCP_cost']))


######################################################################################################################################################

def nat_gateway_cost(dataframe):
  nat_gateway_bytes = dataframe[dataframe[columnnames['usage']].str.contains('NatGateway-Bytes')].copy()
  nat_gateway_hours = dataframe[dataframe[columnnames['usage']].str.contains('NatGateway-Hours')].copy()
  nat_gateway_hours['ninstances'] = nat_gateway_hours[columnnames['quantity']]/720
  ninstances = sum(nat_gateway_hours['ninstances'])
  ngb = sum(nat_gateway_bytes[columnnames['quantity']])
  if ninstances<=32:
    cost = ninstances*nat_gateway_ratecard['us-vm-low']*720+ngb*nat_gateway_ratecard['us-bytes']
  else:
    cost = nat_gateway_ratecard['us-vm-high']*720+ngb*nat_gateway_ratecard['us-bytes']
  return [cost, sum(nat_gateway_bytes[columnnames['cost']])+sum(nat_gateway_hours[columnnames['cost']])]

nat_gcp, nat_aws = nat_gateway_cost(grouped)#df is the grouped dataframe

nat_filter = grouped[columnnames['usage']].str.contains('NatGateway-Bytes') | grouped[columnnames['usage']].str.contains('NatGateway-Hours')
grouped = grouped[~nat_filter]

#feed this input into the viz tool
######################################################################################################################################################

idle_filter = grouped[columnnames['usage']].str.contains('ElasticIP:IdleAddress')
idleaddress = grouped[idle_filter]
grouped = grouped[~idle_filter]

def idle_addresses_cost(row):
  ninstances = int(row[columnnames['quantity']]/720)
  return pd.Series([ninstances*idle_addresses_ratecard['us']*720, row[columnnames['cost']]])

idleaddress[['gcp_cost', 'aws_cost']] = idleaddress.apply(idle_addresses_cost, axis=1)

######################################################################################################################################################

loadbalancer_filter = ((grouped[columnnames['usage']].str.contains('DataProcessing-Bytes'))|(grouped[columnnames['usage']].str.contains('LCUUsage'))|(grouped[columnnames['usage']].str.contains('LoadBalancerUsage')))&(grouped[columnnames['productname']]=='Amazon Elastic Compute Cloud')
loadbalancer = grouped[loadbalancer_filter]
grouped = grouped[~loadbalancer_filter]

def loadbalancer_cost(row):
  rowsplit = row[columnnames['usage']].split('-')
  region_gcp = region_dict[rowsplit[0]]
  usecase = '-'.join(rowsplit[1:])
  return pd.Series([region_gcp, usecase])

if len(loadbalancer):
  loadbalancer[['region', 'usecase']] = loadbalancer.apply(loadbalancer_cost, axis=1)
  load_grouped = loadbalancer.groupby(['region', 'usecase']).sum().reset_index()[['region', 'usecase', columnnames['quantity']]]

def loadgrouped_cost(row):
  region = row['region']
  usecase = row['usecase']
  quantity = row[columnnames['quantity']]
  if usecase=='LoadBalancerUsage':
    nrules = ceil(quantity/720)
    cost = loadbalancer_ratecard['forwarding_rules'][region]*720
    if nrules>5:
      cost+=(nrules-5)*loadbalancer_ratecard['forwarding_rules_extra'][region]*720
    return pd.Series([nrules, cost])
  else:
    cost = quantity*loadbalancer_ratecard['ingress'][region]
    return pd.Series([quantity, cost])
if len(loadbalancer):
  load_grouped[['rules/gb', 'gcp_cost']] = load_grouped.apply(loadgrouped_cost, axis=1)

######################################################################################################################################################

cost_matrix = {
  'aws':{
    'box_compute' : round(sum(boxusage[columnnames['cost']]), 2),
    'heavy_compute' : round(sum(heavyusage[columnnames['cost']]), 2),
    'spot_compute' : round(sum(spotusage[columnnames['cost']]), 2),
    'persistentdisk' : round(sum(persistentdisk[columnnames['cost']]), 2),
    'loadbalancer' : round(sum(loadbalancer[columnnames['cost']]),2),
    'cloudnat' : round(nat_aws, 2),
    'idleaddress' : round(sum(idleaddress['aws_cost']), 2),
  },
  'gcp':{
    'box_compute': 0,
    'heavy_compute' : 0,
    'spot_compute' : 0,
    'persistentdisk' : 0,
    'loadbalancer' : 0,
    'cloudnat' : 0,
    'idleaddress' : 0,

  },
}

if len(boxusage):
  cost_matrix['gcp']['box_usage'] = round(sum(boxusage['gcp_cost']), 2)

if len(heavyusage):
  cost_matrix['gcp']['heavy_usage'] = round(sum(heavyusage['gcp_cost']), 2)

if len(spotusage):
  cost_matrix['gcp']['spot_usage'] = round(sum(spotusage['gcp_cost']), 2)

if len(persistentdisk):
  cost_matrix['gcp']['persistentdisk'] = round(sum(persistentdisk['GCP_cost']), 2)

if len(loadbalancer):
  cost_matrix['gcp']['loadbalancer'] = round(sum(load_grouped['gcp_cost']), 2)

if nat_aws:
  cost_matrix['gcp']['cloudnat'] = round(nat_gcp, 2)

if len(idleaddress):
  cost_matrix['gcp']['idleaddress'] = round(sum(idleaddress['gcp_cost']), 2)

######################################################################################################################################################

wb = Workbook()

report_filename = 'tco_report.xlsx'

ws_summary = wb.active
ws_summary.title = 'Summary'

ws_summary.append([''])
ws_summary.append(['', 'Summarized View - Total Cost of Ownership on GCP'])
center = ws_summary['B2']
center.alignment = Alignment(horizontal='center')
ws_summary.merge_cells(start_row=2, start_column=2, end_row=2, end_column=7)
ws_summary.append(['', 'Compute', 'GCE - Preemptible VMs', '${}'.format(cost_matrix['gcp']['spot_compute']), 'EC2 - Spot Usage', '${}'.format(cost_matrix['aws']['spot_compute'])])
ws_summary.append(['', 'Compute', 'GCE - Regular VMs(Box)', '${}'.format(cost_matrix['gcp']['box_compute']), 'EC2 - Regular Usage(Box)', '${}'.format(cost_matrix['aws']['box_compute'])])
ws_summary.append(['', 'Compute', 'GCE - Regular VMs(Heavy)', '${}'.format(cost_matrix['gcp']['heavy_compute']), 'EC2 - Regular Usage(Heavy)', '${}'.format(cost_matrix['aws']['heavy_compute'])])

ws_box = wb.create_sheet(title='BoxUsage')
box_pyxl = dataframe_to_rows(boxusage)
for r_idx, row in enumerate(box_pyxl, 1):
    for c_idx, value in enumerate(row, 1):
         ws_box.cell(row=r_idx, column=c_idx, value=value)

ws_heavy = wb.create_sheet(title='HeavyUsage')
heavy_pyxl = dataframe_to_rows(heavyusage)
for r_idx, row in enumerate(heavy_pyxl, 1):
    for c_idx, value in enumerate(row, 1):
         ws_heavy.cell(row=r_idx, column=c_idx, value=value)

ws_spot = wb.create_sheet(title='SpotUsage')
spot_pyxl = dataframe_to_rows(spotusage)
for r_idx, row in enumerate(spot_pyxl, 1):
    for c_idx, value in enumerate(row, 1):
         ws_spot.cell(row=r_idx, column=c_idx, value=value)

ws_summary.append(['', '', '', '${}'.format(cost_matrix['gcp']['box_compute']+cost_matrix['gcp']['heavy_compute']+cost_matrix['gcp']['spot_compute']), '', '${}'.format(cost_matrix['aws']['box_compute']+cost_matrix['aws']['heavy_compute']+cost_matrix['aws']['spot_compute'])])

ws_summary.append(['', 'Storage', 'Persistent Disk', '${}'.format(cost_matrix['gcp']['persistentdisk']), 'Elastic Block Storage', '${}'.format(cost_matrix['aws']['persistentdisk'])])
ws_pd = wb.create_sheet(title='Persistent Disk')
pd_pyxl = dataframe_to_rows(persistentdisk)
for r_idx, row in enumerate(pd_pyxl, 1):
    for c_idx, value in enumerate(row, 1):
         ws_pd.cell(row=r_idx, column=c_idx, value=value)

ws_summary.append(['', 'Storage', 'Cloud Storage', '', 'Simple Storage Service(S3)', ''])
ws_summary.append(['', 'Storage', 'Filestore', '', 'Elastic File Storage', ''])
ws_summary.append(['', '', '', '', '', ''])
ws_summary.append(['', 'Networking', 'Cloud Load Balancer', '${}'.format(cost_matrix['gcp']['loadbalancer']), 'Elastic Load Balancer', '${}'.format(cost_matrix['aws']['loadbalancer'])])
ws_summary.append(['', 'Networking', 'Cloud NAT', '${}'.format(cost_matrix['gcp']['cloudnat']), 'NAT Gateway', '${}'.format(cost_matrix['aws']['cloudnat'])])
ws_summary.append(['', 'Networking', 'Network Egress', '', 'Data Transfer', ''])
ws_summary.append(['', 'Networking', 'Idle Addresseses', '${}'.format(cost_matrix['gcp']['idleaddress']), 'Idle Addresses', '${}'.format(cost_matrix['aws']['idleaddress'])])
ws_summary.append(['', '', '', '', '', ''])
ws_summary.append(['', 'DB Services', 'Cloud SQL', '', 'Amazon RDS', ''])
ws_summary.append(['', 'DB Services', 'Search on GCP', '', 'ElasticSearch', ''])
ws_summary.append(['', 'DB Services', 'Cache on GCP', '', 'Elasticache', ''])
ws_summary.append(['', 'DB Services', 'BigQuery', '', 'Redshit', ''])
ws_summary.append(['', '', '', '', '', ''])
ws_summary.append(['', 'Support', 'GCP Support', '', 'AWS Support Business', ''])
ws_summary.append(['', 'Misc', 'Unclassified', '', 'Misc', ''])
ws_summary.append(['', '', '', '', '', ''])
ws_summary.append(['', 'GCP Monthly ', '', 'AWS Monthly', '', ''])

#Throw everything else into the 'Other' tab

ws_others = wb.create_sheet(title='Uncomputed')
others_pyxl = dataframe_to_rows(grouped)
for r_idx, row in enumerate(others_pyxl, 1):
    for c_idx, value in enumerate(row, 1):
         ws_others.cell(row=r_idx, column=c_idx, value=value)


wb.save(filename=report_filename)


######################################################################################################################################################

print('Remaining ', sum(grouped[columnnames['cost']]))